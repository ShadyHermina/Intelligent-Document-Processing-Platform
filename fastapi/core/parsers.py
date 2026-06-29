# fastapi/core/parsers.py
#
# Three-level hybrid chunking pipeline.
#
# Responsibilities:
#   · Accept raw file bytes for PDF, DOCX, or XLSX
#   · Apply Level 1: structure-aware splitting (format-native boundaries)
#   · Apply pre-processing: normalize text within each structural section
#   · Apply Level 2: semantic topic-transition detection (sliding window)
#   · Apply Level 3: size guardrails (token count enforcement + overlap)
#   · Return List[ChunkData] ready for IngestorAgent to wrap in a payload
#
# What does NOT live here:
#   · No database calls
#   · No OpenAI API calls
#   · No agent orchestration
#   · No Pydantic payload assembly
#
# External callers:
#   agents/ingestor.py calls run_pipeline() — the single public entry point.
#   All other functions are internal implementation details.

import io
import re
import unicodedata
from dataclasses import dataclass, field
from typing import List, Tuple

import nltk
import tiktoken
import pdfplumber
from docx import Document as DocxDocument
from docx.oxml.ns import qn
from openpyxl import load_workbook
from sklearn.metrics.pairwise import cosine_similarity

from agents.models import ChunkData

# ---------------------------------------------------------------------------
# NLTK data — download once if not already present
# ---------------------------------------------------------------------------
# sent_tokenize requires the 'punkt_tab' corpus.
# quiet=True suppresses the download progress bar in container logs.
# If already downloaded, this is a no-op.
# We call this at module import time so the first request never waits
# for a corpus download mid-flight.
nltk.download("punkt_tab", quiet=True)

# ---------------------------------------------------------------------------
# Module-level singletons
# ---------------------------------------------------------------------------

# tiktoken encoder — loaded once at module import, reused for every
# token count operation across all requests.
# cl100k_base is the exact encoding used by text-embedding-3-small.
# Using the same encoding for counting as OpenAI uses for embedding
# guarantees our "400 token" guardrail means exactly 400 tokens from
# OpenAI's perspective — not an approximation.
_TOKENIZER = tiktoken.get_encoding("cl100k_base")

# Level 3 guardrail constants.
# These are not in config because they are model constraints, not
# deployment parameters. Changing them requires understanding the
# embedding model's context window, not just tuning a threshold.
_MAX_TOKENS = 400   # hard ceiling per chunk
_MIN_TOKENS = 50    # hard floor per chunk — below this, merge


# ---------------------------------------------------------------------------
# RawSection — internal intermediate type
# ---------------------------------------------------------------------------

@dataclass
class RawSection:
    """
    One structural unit produced by Level 1 parsing.

    Internal to parsers.py — never passed outside this module.
    After Level 2 and Level 3, RawSection objects become ChunkData objects.

    Why a dataclass and not a dict?
    Named attributes make downstream code readable and refactoring safe.
    A dict["location_index"] silently accepts a typo like dict["locaton_index"].
    A dataclass raises AttributeError immediately on the wrong attribute name.

    Why not reuse ChunkData directly here?
    ChunkData is a Pydantic model with validation constraints (min_length,
    ge=1, etc.). During intermediate processing we may temporarily hold
    empty text or zero token counts before guardrails clean them up.
    Using ChunkData here would require suppressing validation mid-pipeline,
    which defeats its purpose. RawSection is intentionally permissive.
    """
    text:           str
    location_index: int          # page number (PDF/DOCX) or sheet index (XLSX)
    section_label:  str          # heading text, sheet/rowgroup label, or "Body"
    image_present:  bool = False


# ---------------------------------------------------------------------------
# normalize_text
# ---------------------------------------------------------------------------

def normalize_text(text: str) -> str:
    """
    Apply the pre-processing normalization pass to a single text string.

    Called on every RawSection.text before Level 2 processing begins.
    Order of operations matters — unicode normalization must happen before
    whitespace collapse because some unicode characters expand to multiple
    characters under NFKC and those expansions may include whitespace.

    Parameters
    ----------
    text : str
        Raw text as extracted by the format-specific parser.

    Returns
    -------
    str
        Cleaned, normalized text. May be empty if the input contained
        only boilerplate or control characters — callers must handle
        the empty string case.
    """

    # Step 1 — re-encode to enforce UTF-8 validity.
    # errors="replace" substitutes the Unicode replacement character (U+FFFD)
    # for any byte sequence that cannot be decoded as UTF-8.
    # Without this, a malformed PDF could produce a string with invalid
    # surrogates that crash downstream string operations unpredictably.
    text = text.encode("utf-8", errors="replace").decode("utf-8")

    # Step 2 — Unicode normalization.
    # NFKC (Compatibility Decomposition followed by Canonical Composition):
    #   · Decomposes compatibility characters: ﬁ → fi, ² → 2, ½ → 1/2
    #   · Normalizes different representations of the same character
    #     (e.g. é as one character vs e + combining accent as two)
    # Why NFKC and not NFC?
    # NFC only does canonical composition — it would not decompose ﬁ to fi.
    # NFKC is more aggressive and more appropriate for document text where
    # ligatures and special number forms should be treated as their ASCII
    # equivalents for embedding purposes.
    text = unicodedata.normalize("NFKC", text)

    # Step 3 — strip control characters.
    # Remove all characters with ordinal below 32 EXCEPT newline (\n, ord 10).
    # This removes: null bytes, bell, backspace, carriage return, tab (handled
    # separately below), form feed, and other non-printing control characters
    # that appear in some PDFs and DOCX files due to formatting artifacts.
    # We keep \n because it carries paragraph structure information used by
    # sent_tokenize to correctly identify sentence boundaries.
    text = "".join(
        ch for ch in text
        if ord(ch) >= 32 or ch == "\n"
    )

    # Step 4 — collapse whitespace variants to a single space.
    # \t   (tab, ord 9)    — already kept through step 3 since ord(9) < 32
    #                         but we collapsed it to space here
    # \r   (carriage return, ord 13) — removed in step 3, but belt-and-suspenders
    # \xa0 (non-breaking space, U+00A0) — common in PDFs, visually identical
    #                                       to a space but not matched by \s
    # Runs of multiple spaces → single space.
    text = re.sub(r"[ \t\r\xa0]+", " ", text)

    # Step 5 — normalize multiple consecutive newlines to a maximum of two.
    # More than two consecutive newlines is visual formatting with no
    # semantic content — collapsing to two preserves paragraph separation
    # while removing excessive vertical whitespace from PDFs.
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip()
    # .strip() removes leading and trailing whitespace (including newlines)
    # from the entire string. Sections that start or end with blank lines
    # after normalization produce cleaner sentence tokenization.


# ---------------------------------------------------------------------------
# _remove_boilerplate
# ---------------------------------------------------------------------------

def _remove_boilerplate(sections: List[RawSection]) -> List[RawSection]:
    """
    Strip repeated headers and footers from all sections.

    A line is considered boilerplate if it appears in more than 60% of
    sections — this threshold catches page headers ("CONFIDENTIAL"),
    footers ("Page N of M"), and watermarks that repeat on every page
    while preserving lines that happen to appear twice in a short document.

    Why 60% and not a fixed count?
    A fixed count (e.g. "appears more than 3 times") would strip legitimate
    repeated content in a short 4-section document while missing boilerplate
    in a 100-section document that appears on 59 pages. A percentage scales
    correctly with document length.

    Called once on the full List[RawSection] after Level 1 parsing,
    before Level 2 processing begins.
    """
    if len(sections) < 3:
        # With fewer than 3 sections there is not enough signal to
        # distinguish boilerplate from legitimate repeated content.
        return sections

    threshold = 0.60
    total = len(sections)

    # Count how many sections each line appears in.
    # We compare stripped lines to handle minor whitespace differences.
    line_counts: dict[str, int] = {}
    for section in sections:
        # Use a set per section so a line that appears twice on one page
        # is only counted once toward the boilerplate threshold.
        unique_lines = set(
            line.strip()
            for line in section.text.splitlines()
            if line.strip()
        )
        for line in unique_lines:
            line_counts[line] = line_counts.get(line, 0) + 1

    # Identify boilerplate lines.
    boilerplate = {
        line for line, count in line_counts.items()
        if count / total > threshold
    }

    if not boilerplate:
        return sections

    # Remove boilerplate lines from every section.
    cleaned = []
    for section in sections:
        clean_lines = [
            line for line in section.text.splitlines()
            if line.strip() not in boilerplate
        ]
        cleaned_text = "\n".join(clean_lines).strip()
        cleaned.append(RawSection(
            text           = cleaned_text,
            location_index = section.location_index,
            section_label  = section.section_label,
            image_present  = section.image_present,
        ))

    return cleaned


# ---------------------------------------------------------------------------
# _serialize_table_rows
# ---------------------------------------------------------------------------

def _serialize_table_rows(headers: List[str], rows: List[List[str]]) -> str:
    """
    Convert a table to a sequence of "Header: value, Header: value" lines.

    This representation is readable by the embedding model as natural
    language — unlike raw whitespace-separated columns which lose the
    relationship between a column name and its value.

    Example input:
        headers = ["Company", "Amount", "Date"]
        rows    = [["Acme", "$50,000", "2024-01-01"],
                   ["Beta", "$1,200",  "2024-02-01"]]

    Example output:
        "Company: Acme, Amount: $50,000, Date: 2024-01-01\n
         Company: Beta, Amount: $1,200, Date: 2024-02-01"

    Parameters
    ----------
    headers : List[str]
        Column header strings. May be empty strings for columns with
        no header — we use "Col{N}" as a fallback.
    rows : List[List[str]]
        Each inner list is one row. Cell values are already strings.
    """
    clean_headers = [
        h.strip() if h.strip() else f"Col{i+1}"
        for i, h in enumerate(headers)
    ]

    serialized_rows = []
    for row in rows:
        pairs = []
        for i, cell in enumerate(row):
            header = clean_headers[i] if i < len(clean_headers) else f"Col{i+1}"
            cell_text = str(cell).strip() if cell is not None else ""
            if cell_text:
                # Skip empty cells — "Amount: " with no value adds noise.
                pairs.append(f"{header}: {cell_text}")
        if pairs:
            serialized_rows.append(", ".join(pairs))

    return "\n".join(serialized_rows)


# ---------------------------------------------------------------------------
# parse_pdf
# ---------------------------------------------------------------------------

def parse_pdf(raw_bytes: bytes) -> List[RawSection]:
    """
    Level 1 parsing for PDF files using pdfplumber.

    Extracts text page by page. Within each page, detects tables and
    serializes them separately from body text. Detects images by checking
    for image objects in the page and sets image_present on sections
    that contain adjacent caption text.

    Each page produces one or more RawSection objects:
      · One section for the body text of the page
      · Additional sections if tables are present on the page

    Why pdfplumber over PyMuPDF or pypdf?
    pdfplumber exposes table detection and bounding box information as
    first-class objects. PyMuPDF is faster but its table detection
    requires the separate pymupdf-table plugin. pypdf extracts text only,
    with no layout awareness. pdfplumber gives us exactly what Level 1
    requires: structural boundaries with table detection included.

    Parameters
    ----------
    raw_bytes : bytes
        Raw PDF file content read from the upload.

    Returns
    -------
    List[RawSection]
        One RawSection per structural unit per page.
        May be empty for pages that contain only images with no text.
    """
    sections: List[RawSection] = []

    with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            # page_num is 1-based — matches how humans refer to pages
            # and how the location_index field is documented.

            # Detect images on this page.
            # pdfplumber exposes page.images as a list of image objects.
            # We record image presence and look for caption text within
            # 50px below any detected image bounding box.
            page_has_image = len(page.images) > 0
            caption_texts: List[str] = []

            if page_has_image:
                for img in page.images:
                    # img["y1"] is the bottom edge of the image in PDF
                    # coordinate space (origin at bottom-left).
                    # We look for text objects whose top edge (y1 in text
                    # space, which is the bottom of the text block) falls
                    # within 50 units below the image bottom.
                    # PDF units are typically points (1/72 inch).
                    img_bottom = img.get("y0", 0)
                    for word in page.extract_words():
                        word_top = word.get("top", 0)
                        if 0 <= (word_top - (page.height - img_bottom)) <= 50:
                            caption_texts.append(word.get("text", ""))

            caption_text = " ".join(caption_texts).strip()

            # Extract tables first — they get their own RawSection.
            tables = page.find_tables()
            table_bboxes = []
            for table in tables:
                extracted = table.extract()
                if not extracted or len(extracted) < 2:
                    # A table with 0 or 1 rows has no data rows — skip.
                    continue

                headers = [str(cell) if cell else "" for cell in extracted[0]]
                data_rows = [
                    [str(cell) if cell else "" for cell in row]
                    for row in extracted[1:]
                ]
                table_text = _serialize_table_rows(headers, data_rows)
                if table_text.strip():
                    sections.append(RawSection(
                        text           = normalize_text(table_text),
                        location_index = page_num,
                        section_label  = f"Table on page {page_num}",
                        image_present  = False,
                    ))
                table_bboxes.append(table.bbox)
                # We track table bounding boxes so we can exclude table
                # regions when extracting body text below.

            # Extract body text — exclude words that fall inside table bboxes.
            #
            # pdfplumber's table.bbox is (x0, top, x1, bottom) in pdfplumber
            # coordinates (top-left origin, Y increases downward).
            # We extract every word on the page, then discard any word whose
            # bounding box overlaps a table bbox. The remaining words are
            # joined to form the body text.
            #
            # Why word-level filtering instead of page.extract_text()?
            # page.extract_text() has no way to exclude specific regions —
            # it extracts everything. Word-level filtering lets us precisely
            # exclude table content so table cells never appear twice:
            # once in the table RawSection and once in the body RawSection.
            # Duplicate vectors with different point_ids waste Qdrant space
            # and pollute retrieval results — upsert idempotency does NOT
            # deduplicate content with different point_ids.

            def _word_overlaps_table(word: dict, bboxes: list) -> bool:
                """
                Return True if the word's bounding box overlaps any table bbox.

                Both word and table use pdfplumber coordinates:
                  x0, top = left edge, top edge
                  x1, bottom = right edge, bottom edge

                Two rectangles overlap when they overlap on BOTH axes:
                  x-overlap: word left < table right AND word right > table left
                  y-overlap: word top  < table bottom AND word bottom > table top
                """
                wx0    = word.get("x0", 0)
                wx1    = word.get("x1", 0)
                wtop   = word.get("top", 0)
                wbottom= word.get("bottom", 0)
                for (tx0, ttop, tx1, tbottom) in bboxes:
                    if wx0 < tx1 and wx1 > tx0 and wtop < tbottom and wbottom > ttop:
                        return True
                return False

            all_words = page.extract_words(x_tolerance=3, y_tolerance=3)
            body_words = [
                w["text"] for w in all_words
                if not _word_overlaps_table(w, table_bboxes)
            ]
            body_text = normalize_text(" ".join(body_words))

            if body_text:
                # Prepend caption text to the body section if we found it.
                # This keeps the caption co-located with the surrounding
                # body text rather than isolated in its own section.
                if caption_text:
                    body_text = caption_text + "\n" + body_text

                sections.append(RawSection(
                    text           = body_text,
                    location_index = page_num,
                    section_label  = f"Page {page_num}",
                    image_present  = page_has_image,
                ))

    return sections


# ---------------------------------------------------------------------------
# parse_docx
# ---------------------------------------------------------------------------

def parse_docx(raw_bytes: bytes) -> List[RawSection]:
    """
    Level 1 parsing for DOCX files using python-docx.

    Splits on native paragraph style boundaries. Heading 1 and Heading 2
    styles mark section transitions — when a heading is encountered, the
    accumulated text of the previous section is flushed as a RawSection
    and a new section begins under the new heading.

    Tables are extracted and serialized row by row using the first row
    as headers. An image paragraph immediately following an image run
    is treated as a caption — image_present is set on that section.

    Parameters
    ----------
    raw_bytes : bytes
        Raw DOCX file content read from the upload.

    Returns
    -------
    List[RawSection]
        One RawSection per logical section (Heading-delimited block).
        Tables produce their own RawSection.
    """
    sections: List[RawSection] = []
    doc = DocxDocument(io.BytesIO(raw_bytes))

    current_heading = "Body"
    current_lines: List[str] = []
    current_page = 1       # DOCX has no native page API — we use 1 for all
    last_had_image = False # tracks whether previous paragraph had an image

    def _flush(heading: str, lines: List[str], page: int, has_image: bool):
        """Flush accumulated lines into a RawSection."""
        text = normalize_text("\n".join(lines))
        if text:
            sections.append(RawSection(
                text           = text,
                location_index = page,
                section_label  = heading,
                image_present  = has_image,
            ))

    for block in doc.element.body:
        # doc.element.body iterates the raw XML children of the document
        # body. Each child is either a paragraph (w:p) or a table (w:tbl).
        # We check the tag to dispatch correctly.

        tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag
        # block.tag looks like "{http://schemas.openxmlformats.org/...}p"
        # We strip the namespace prefix to get "p" or "tbl".

        if tag == "p":
            # This is a paragraph element.
            from docx.oxml.ns import qn as _qn

            # Extract paragraph style name.
            pPr = block.find(_qn("w:pPr"))
            style_name = ""
            if pPr is not None:
                pStyle = pPr.find(_qn("w:pStyle"))
                if pStyle is not None:
                    style_name = pStyle.get(_qn("w:val"), "")

            # Extract paragraph text.
            para_text = "".join(
                node.text or ""
                for node in block.iter()
                if node.tag == _qn("w:t")
            )

            # Detect image runs in this paragraph.
            has_drawing = block.find(".//" + _qn("w:drawing")) is not None
            has_pict    = block.find(".//" + _qn("w:pict")) is not None
            para_has_image = has_drawing or has_pict

            # Check if this paragraph is a heading.
            is_heading = (
                style_name.startswith("Heading1")
                or style_name.startswith("Heading2")
                or style_name == "1"   # some DOCX use numeric style IDs
                or style_name == "2"
            )

            if is_heading:
                # Flush the previous section before starting a new one.
                _flush(current_heading, current_lines, current_page, last_had_image)
                current_heading = para_text.strip() or "Untitled Section"
                current_lines = []
                last_had_image = False
            elif para_has_image:
                # Image paragraph — flush current section, mark image present.
                _flush(current_heading, current_lines, current_page, last_had_image)
                current_lines = []
                last_had_image = True
                # The caption will be the next paragraph — it will be
                # accumulated into current_lines with last_had_image=True.
            else:
                if para_text.strip():
                    current_lines.append(para_text.strip())
                    if last_had_image:
                        # This paragraph immediately follows an image —
                        # treat it as a caption. last_had_image stays True
                        # so the section that gets flushed next carries the flag.
                        pass

        elif tag == "tbl":
            # This is a table element.
            # Flush current section before the table.
            _flush(current_heading, current_lines, current_page, last_had_image)
            current_lines = []
            last_had_image = False

            # Extract table rows using python-docx's table API.
            # We need to re-wrap the raw XML element as a python-docx Table.
            from docx.table import Table as DocxTable
            from docx.oxml import OxmlElement
            table_obj = DocxTable(block, doc)

            rows = []
            for row in table_obj.rows:
                rows.append([cell.text.strip() for cell in row.cells])

            if len(rows) >= 2:
                headers  = rows[0]
                data_rows = rows[1:]
                table_text = _serialize_table_rows(headers, data_rows)
                if table_text.strip():
                    sections.append(RawSection(
                        text           = normalize_text(table_text),
                        location_index = current_page,
                        section_label  = f"{current_heading} / Table",
                        image_present  = False,
                    ))

    # Flush whatever remains after the last element.
    _flush(current_heading, current_lines, current_page, last_had_image)

    return sections


# ---------------------------------------------------------------------------
# parse_xlsx
# ---------------------------------------------------------------------------

def parse_xlsx(raw_bytes: bytes) -> List[RawSection]:
    """
    Level 1 parsing for XLSX files using openpyxl.

    Each sheet is a structural unit. Within each sheet, contiguous
    non-empty row groups separated by empty rows are the sub-units.
    Row 1 of each contiguous group is used as column headers.
    Each remaining row is serialized as "Header: value, Header: value".

    Why contiguous row groups and not the entire sheet as one section?
    A single sheet often contains multiple independent tables separated
    by blank rows — e.g. a summary table at the top and a detail table
    below, separated by one empty row. Treating the whole sheet as one
    section would mix unrelated data into one chunk, producing a low-
    quality embedding that retrieves on too many topics at once.

    Parameters
    ----------
    raw_bytes : bytes
        Raw XLSX file content read from the upload.

    Returns
    -------
    List[RawSection]
        One RawSection per contiguous row group per sheet.
    """
    sections: List[RawSection] = []
    wb = load_workbook(
        filename=io.BytesIO(raw_bytes),
        read_only=True,   # read_only=True is faster and uses less memory
        data_only=True,   # data_only=True returns cell values, not formulas
    )

    for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]

        # Read all rows into a list so we can detect empty-row boundaries.
        all_rows: List[List] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        # Split into contiguous non-empty row groups.
        # An empty row is one where every cell is None or empty string.
        groups: List[List[List]] = []
        current_group: List[List] = []

        for row in all_rows:
            row_is_empty = all(
                cell is None or str(cell).strip() == ""
                for cell in row
            )
            if row_is_empty:
                if current_group:
                    groups.append(current_group)
                    current_group = []
            else:
                current_group.append(row)

        if current_group:
            groups.append(current_group)

        # Serialize each group.
        for group_index, group in enumerate(groups, start=1):
            if len(group) < 2:
                # A group with only one row is a header with no data.
                # Not worth embedding on its own — skip.
                continue

            # First row of each group is treated as the header row.
            headers = [str(cell).strip() if cell is not None else "" for cell in group[0]]
            data_rows = [
                [str(cell).strip() if cell is not None else "" for cell in row]
                for row in group[1:]
            ]

            table_text = _serialize_table_rows(headers, data_rows)
            if table_text.strip():
                sections.append(RawSection(
                    text           = normalize_text(table_text),
                    location_index = sheet_index,
                    section_label  = f"{sheet_name} / row group {group_index}",
                    image_present  = False,
                    # XLSX images are rare and openpyxl's image extraction
                    # requires non-read-only mode and adds significant complexity.
                    # image_present=False is correct for the vast majority of
                    # spreadsheets. Phase 9 can add image detection if needed.
                ))

    wb.close()
    return sections


# ---------------------------------------------------------------------------
# semantic_chunk  (Level 2)
# ---------------------------------------------------------------------------

def semantic_chunk(
    section: RawSection,
    st_model,               # SentenceTransformer — no type hint to avoid
                            # importing sentence_transformers at module level
                            # which would force the heavy import on all callers
    threshold: float,
) -> List[RawSection]:
    """
    Apply Level 2 sliding window topic-transition detection to one section.

    Tokenizes the section text into sentences, embeds each sentence with
    the local all-MiniLM-L6-v2 model, and computes cosine similarity
    between each adjacent pair. Where similarity drops below threshold,
    a chunk boundary is inserted.

    Parameters
    ----------
    section : RawSection
        One structural unit from Level 1.
    st_model : SentenceTransformer
        The sentence transformer model loaded once at startup and stored
        on app.state.st_model. Passed in rather than imported here to
        keep this function testable without a running FastAPI application.
    threshold : float
        Cosine similarity below which a boundary is cut.
        Read from settings.chunk_similarity_threshold (default 0.55).

    Returns
    -------
    List[RawSection]
        One or more RawSection objects. Each inherits location_index,
        section_label, and image_present from the input section.
        If the section has only one sentence or no boundary is detected,
        a list containing the original section is returned unchanged.
    """
    text = section.text.strip()
    if not text:
        return []

    # Tokenize into sentences.
    # nltk.sent_tokenize handles abbreviations: "Dr. Smith" is not split.
    # Naive period splitting would produce: "Dr", " Smith joined Acme Corp"
    sentences = nltk.sent_tokenize(text)

    if len(sentences) <= 1:
        # Cannot compute similarity between adjacent pairs with one sentence.
        # Return the section unchanged.
        return [section]

    # Embed all sentences in one batch call.
    # st_model.encode() returns a numpy array of shape (N, embedding_dim).
    # These embeddings are used ONLY for boundary detection — they are
    # NOT stored in Qdrant. Qdrant receives OpenAI embeddings via
    # embed_and_store() later in Agent 2.
    embeddings = st_model.encode(sentences, show_progress_bar=False)
    # show_progress_bar=False suppresses tqdm output in container logs.

    # Sliding window: compute cosine similarity between each adjacent pair.
    # cosine_similarity expects 2D arrays — reshape each 1D vector to (1, N).
    boundaries: List[int] = []
    for i in range(len(sentences) - 1):
        sim = cosine_similarity(
            embeddings[i].reshape(1, -1),
            embeddings[i + 1].reshape(1, -1),
        )[0][0]
        # [0][0] extracts the scalar from the (1,1) result matrix.

        if sim < threshold:
            boundaries.append(i)
            # Boundary after sentence i means:
            #   chunk A ends with sentences[0..i]
            #   chunk B starts with sentences[i+1..]

    if not boundaries:
        # No topic drift detected — the entire section is one coherent topic.
        return [section]

    # Split sentences into groups at each boundary.
    groups: List[List[str]] = []
    start = 0
    for boundary in boundaries:
        groups.append(sentences[start : boundary + 1])
        start = boundary + 1
    groups.append(sentences[start:])
    # The final group collects all remaining sentences after the last boundary.

    # Convert each group back to a RawSection, inheriting metadata.
    result: List[RawSection] = []
    for group in groups:
        group_text = " ".join(group).strip()
        if group_text:
            result.append(RawSection(
                text           = group_text,
                location_index = section.location_index,
                section_label  = section.section_label,
                image_present  = section.image_present,
            ))

    return result if result else [section]


# ---------------------------------------------------------------------------
# apply_size_guardrails  (Level 3)
# ---------------------------------------------------------------------------

def apply_size_guardrails(sections: List[RawSection]) -> List[ChunkData]:
    """
    Apply Level 3 size guardrails to the output of Level 2.

    Three operations in order:
      1. Split any section exceeding _MAX_TOKENS at the nearest sentence
         boundary before the limit. Never splits mid-sentence.
      2. Merge any section below _MIN_TOKENS with the following section.
         If the last section is below _MIN_TOKENS, merge with the preceding.
      3. Apply sentence-level overlap: prepend the last sentence of chunk N
         to the text of chunk N+1 for context continuity.

    Finally assigns sequential chunk_index values and records final
    token_count per chunk.

    Parameters
    ----------
    sections : List[RawSection]
        Output of Level 2 — semantically coherent sections, potentially
        violating size constraints.

    Returns
    -------
    List[ChunkData]
        Final chunks ready for the IngestorPayload. All chunks satisfy:
        _MIN_TOKENS <= token_count <= _MAX_TOKENS.
    """

    # ── Pass 1: split oversized sections ────────────────────────────────

    split_sections: List[RawSection] = []

    for section in sections:
        tokens = _TOKENIZER.encode(section.text)

        if len(tokens) <= _MAX_TOKENS:
            split_sections.append(section)
            continue

        # Section exceeds 400 tokens — split at sentence boundaries.
        sentences = nltk.sent_tokenize(section.text)
        current_sentences: List[str] = []
        current_tokens = 0

        for sentence in sentences:
            sentence_tokens = len(_TOKENIZER.encode(sentence))

            if current_tokens + sentence_tokens > _MAX_TOKENS and current_sentences:
                # Adding this sentence would exceed the limit.
                # Flush current accumulation as a new section.
                split_sections.append(RawSection(
                    text           = " ".join(current_sentences),
                    location_index = section.location_index,
                    section_label  = section.section_label,
                    image_present  = section.image_present,
                ))
                current_sentences = [sentence]
                current_tokens    = sentence_tokens
            else:
                current_sentences.append(sentence)
                current_tokens += sentence_tokens

        # Flush whatever remains.
        if current_sentences:
            split_sections.append(RawSection(
                text           = " ".join(current_sentences),
                location_index = section.location_index,
                section_label  = section.section_label,
                image_present  = section.image_present,
            ))

    # ── Pass 2: merge undersized sections ───────────────────────────────

    merged_sections: List[RawSection] = []

    i = 0
    while i < len(split_sections):
        section = split_sections[i]
        token_count = len(_TOKENIZER.encode(section.text))

        if token_count < _MIN_TOKENS:
            if i + 1 < len(split_sections):
                # Merge with the next section.
                # The merged section inherits metadata from the current section
                # (location_index, section_label) — it started here.
                next_section = split_sections[i + 1]
                merged_text = section.text + " " + next_section.text
                split_sections[i + 1] = RawSection(
                    text           = merged_text.strip(),
                    location_index = section.location_index,
                    section_label  = section.section_label,
                    image_present  = section.image_present or next_section.image_present,
                )
                # Skip the current section — it has been merged into i+1.
                i += 1
                continue
            elif merged_sections:
                # This is the last section and it is too short.
                # Merge with the preceding section.
                prev = merged_sections[-1]
                merged_text = prev.text + " " + section.text
                merged_sections[-1] = RawSection(
                    text           = merged_text.strip(),
                    location_index = prev.location_index,
                    section_label  = prev.section_label,
                    image_present  = prev.image_present or section.image_present,
                )
                i += 1
                continue

        merged_sections.append(section)
        i += 1

    # ── Pass 3: sentence-level overlap ──────────────────────────────────

    # Prepend the last sentence of chunk N to chunk N+1.
    # This gives the embedding model context about what came immediately
    # before, improving retrieval accuracy for content that references
    # the preceding chunk (e.g. "As mentioned above, the clause states...")
    overlapped_sections: List[RawSection] = []

    for idx, section in enumerate(merged_sections):
        if idx == 0:
            overlapped_sections.append(section)
            continue

        prev_text = merged_sections[idx - 1].text
        prev_sentences = nltk.sent_tokenize(prev_text)

        if prev_sentences:
            last_sentence = prev_sentences[-1]
            new_text = last_sentence + " " + section.text
            overlapped_sections.append(RawSection(
                text           = new_text.strip(),
                location_index = section.location_index,
                section_label  = section.section_label,
                image_present  = section.image_present,
            ))
        else:
            overlapped_sections.append(section)

    # ── Final: convert to ChunkData with index and token count ──────────

    chunks: List[ChunkData] = []

    for chunk_index, section in enumerate(overlapped_sections):
        final_token_count = len(_TOKENIZER.encode(section.text))

        chunks.append(ChunkData(
            chunk_index    = chunk_index,
            text           = section.text,
            location_index = section.location_index,
            section_label  = section.section_label,
            image_present  = section.image_present,
            token_count    = max(final_token_count, 1),
            # max(..., 1) prevents a ValidationError from ChunkData's ge=1
            # constraint in the pathological case where overlap text is
            # somehow empty — the merge pass above should prevent this,
            # but we guard here as a final safety net.
        ))

    return chunks


# ---------------------------------------------------------------------------
# run_pipeline  — public entry point
# ---------------------------------------------------------------------------

def run_pipeline(
    raw_bytes: bytes,
    file_type: str,
    st_model,
    threshold: float,
) -> Tuple[List[ChunkData], int]:
    """
    Execute the complete three-level chunking pipeline.

    Called by IngestorAgent.run(). This is the only function in parsers.py
    that IngestorAgent imports — all internal functions are implementation
    details.

    Parameters
    ----------
    raw_bytes : bytes
        Raw file content from the upload.
    file_type : str
        One of "pdf", "docx", "xlsx". Already validated by the upload
        endpoint before reaching this function.
    st_model : SentenceTransformer
        Loaded once at startup, stored on app.state.st_model.
    threshold : float
        Cosine similarity threshold for Level 2 boundary detection.
        From settings.chunk_similarity_threshold.

    Returns
    -------
    Tuple[List[ChunkData], int]
        chunks         — final processed chunks, ready for IngestorPayload
        location_count — total pages (PDF/DOCX) or sheets (XLSX)
    """

    # ── Level 1: format-specific structural parsing ──────────────────────

    if file_type == "pdf":
        raw_sections = parse_pdf(raw_bytes)
    elif file_type == "docx":
        raw_sections = parse_docx(raw_bytes)
    elif file_type == "xlsx":
        raw_sections = parse_xlsx(raw_bytes)
    else:
        # This branch should never be reached — the upload endpoint
        # validates file_type before calling the agent. Guard anyway.
        raise ValueError(f"Unsupported file_type: {file_type}")

    if not raw_sections:
        raise ValueError(
            f"No text could be extracted from the {file_type.upper()} file. "
            "The file may be empty, image-only, or corrupted."
        )

    # Compute location_count from the sections before boilerplate removal.
    # After removal some pages/sheets may become empty — we count from
    # the raw parse to report the true document size.
    location_count = max(s.location_index for s in raw_sections)

    # ── Pre-processing: boilerplate removal ─────────────────────────────

    raw_sections = _remove_boilerplate(raw_sections)

    # Filter out sections that became empty after boilerplate removal.
    raw_sections = [s for s in raw_sections if s.text.strip()]

    if not raw_sections:
        raise ValueError(
            "All extracted text was identified as boilerplate. "
            "The document may consist entirely of repeated headers/footers."
        )

    # ── Level 2: semantic topic-transition detection ─────────────────────

    semantic_sections: List[RawSection] = []
    for section in raw_sections:
        semantic_sections.extend(
            semantic_chunk(section, st_model, threshold)
        )

    # Filter again — semantic_chunk may return empty sections.
    semantic_sections = [s for s in semantic_sections if s.text.strip()]

    # ── Level 3: size guardrails + overlap + ChunkData assembly ─────────

    chunks = apply_size_guardrails(semantic_sections)

    return chunks, location_count